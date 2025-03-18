import os
os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ["CUDA_VISIBLE_DEVICES"] = "0,1,2,3" # Change this to your GPU IDs
import torch
import gc
import argparse
import pandas as pd
from transformers.cache_utils import DynamicCache
import time
import threading
import queue
import re
import multiprocessing
from transformers import (
    AutoModelForCausalLM,
    AutoTokenizer
)
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook
from peft import PeftModel
from util import (
  readOpenROADOutput,
  runOpenROADShell,
  sendCommandOpenROAD,
  processCodeString,
  clearQueue,
  generate,
  modelUtility,
  prepareDocuments,
  answerWithRAG
)

def Run(
  testSetPath: str,
  resultPath: str,
  RAGApiPath: str,
  RAGCodePath: str,
  modelName: str,
  OpenROADPath: str,
  loadDesignTime: float,
  maxTestCaseWaitTime: float,
  commandFlushTime: float
):
  crossStageLoadDesignCommand = ["file", "file", "floorplan", "floorplan", 
    "floorplan", "floorplan", "io", "io", "gpl", "gpl", "dpl", "dpl",
    "dpl", "pdn", "grt", "file", "file", "file", "file", "file"
  ]

  promptSet = pd.read_excel(testSetPath, 'Prompt', header=None)
  promptSet = promptSet.rename(columns={0: "0"})

  correctCodeSet = pd.read_excel(testSetPath, 'Code', header=None)
  correctCodeSet = correctCodeSet.rename(columns={0: "0"})
  codeTestSet = Workbook()

  codeTestSetIter = list()
  for i in range(6):
    codeTestSetIter.append(codeTestSet.create_sheet("Sheet" + str(i)))
  codeTestSetIter.append(codeTestSet.create_sheet("Flow"))

  for i in range(7):
    codeTestSetIter[i]["A1"] = "correct code" 
    codeTestSetIter[i]["B1"] = "prompt"
    codeTestSetIter[i]["C1"] = "code1"
    codeTestSetIter[i]["D1"] = "output1"
    codeTestSetIter[i]["E1"] = "code2"
    codeTestSetIter[i]["F1"] = "output2"        
    codeTestSetIter[i]["G1"] = "code3"
    codeTestSetIter[i]["H1"] = "output3"
  
  index = 1
  for i in range(120):
    if i%6 == 0:
      index += 1
    codeTestSetIter[int(i%6)]["A"+str(index)] = correctCodeSet["0"][i]
    codeTestSetIter[int(i%6)]["B"+str(index)] = promptSet["0"][i]

  for i in range(20):
    codeTestSetIter[6]["A"+str(i+2)] = correctCodeSet["0"][i+120]
    codeTestSetIter[6]["B"+str(i+2)] = promptSet["0"][i+120]

  gc.collect()

  apiDf = pd.read_csv(RAGApiPath)
  apiDocuments, apiDocumentsDict = prepareDocuments(df=apiDf)
  templateDf = pd.read_csv(RAGCodePath)
  templateDocuments, templateDocumentsDict = prepareDocuments(df=templateDf, api=False)
  allSplits = apiDocuments + templateDocuments
  allDict = {**apiDocumentsDict, **templateDocumentsDict}

  embeddingModel = SentenceTransformer("mixedbread-ai/mxbai-embed-large-v1")
  embeddings = embeddingModel.encode(allSplits)

  util = modelUtility(modelName)

  OpenROADProcess = None
  tokenizer = None
  model = None

  if "script_adaptor" in modelName.lower() or "llama" in modelName.lower() or "retrained" in modelName.lower():
    tokenizer = AutoTokenizer.from_pretrained(
      modelName,
      pad_token = '<|end_of_text|>',
      eos_token = '<|eot_id|>',
      cache_dir = None,
      truncation = True,
      padding_side = "right",
      trust_remote_code = True,
      device_map = "balanced_low_0"
    )
    tokenizer.add_special_tokens({"additional_special_tokens": ["<|eot_id|>", "<|end_of_text|>"]})
  else:
    tokenizer = AutoTokenizer.from_pretrained(
      modelName,
      pad_token = '<|endoftext|>',
      eos_token = '<|im_end|>',
      cache_dir = None,
      truncation = True,
      padding_side = "right",
      trust_remote_code = True,
      device_map = "balanced_low_0"
    )
    tokenizer.add_special_tokens({"additional_special_tokens": ["<|im_end|>", "<|im_start|>"]})

  if "retrained" in modelName.lower():
    model = AutoModelForCausalLM.from_pretrained(
      "meta-llama/Meta-Llama-3-8B-Instruct",
      device_map="balanced_low_0",
      torch_dtype=torch.bfloat16,
      attn_implementation="flash_attention_2",
    )
    model = PeftModel.from_pretrained(
      model,
      modelName,
      is_trainable=False,
      device_map="balanced_low_0"
    )
  elif "32b" in modelName.lower() and "agent" in modelName.lower():
    model = AutoModelForCausalLM.from_pretrained(
      "Qwen/Qwen2.5-Coder-32B-Instruct",
      device_map="balanced_low_0",
      torch_dtype=torch.bfloat16,
      attn_implementation="flash_attention_2",
    )
    model = PeftModel.from_pretrained(
      model,
      modelName,
      is_trainable=False,
      device_map="balanced_low_0"
    )
  elif "7b" in modelName.lower() and "agent" in modelName.lower():
    model = AutoModelForCausalLM.from_pretrained(
      "Qwen/Qwen2.5-Coder-7B-Instruct",
      device_map="balanced_low_0",
      torch_dtype=torch.bfloat16,
      attn_implementation="flash_attention_2",
    ) 
    model = PeftModel.from_pretrained(
      model,
      modelName,
      is_trainable=False,
      device_map="balanced_low_0"
    )
  else:
    model = AutoModelForCausalLM.from_pretrained(
      modelName,
      device_map="balanced_low_0",
      torch_dtype=torch.bfloat16,
      attn_implementation="flash_attention_2",
    )

  for sheetIndex in range(7):
    print("==================="+str(sheetIndex)+"===================")
    for i in range(20):
      print("==================="+str(sheetIndex)+"-"+str(i)+"===================")
      for passIndex in ["C", "E", "G"]:
        prompt = codeTestSetIter[sheetIndex]["B" + str(i + 2)].value
        RAGContext = answerWithRAG(
          prompt,
          embeddings,
          embeddingModel,
          allSplits,
          allDict
        )
        if RAGContext == "":
          prompt = util.ragPromptTemplateWithoutContext.format(
            question = prompt,
            system_prompt = util.systemPrompt
          )
        else:
          prompt = util.ragPromptTemplateWithContext.format(
            question = prompt,
            context = RAGContext,
            system_prompt = util.systemPrompt
          )
        
        if OpenROADProcess is not None:
            OpenROADProcess.terminate()
            OpenROADProcess.wait()
        # Start the OpenROAD process
        loadDesignType = "" if sheetIndex < 6 else crossStageLoadDesignCommand[i]
        print(loadDesignType)
        OpenROADProcess = runOpenROADShell(OpenROADPath, loadDesignTime, slaveOpenROAD, loadDesignType)
        time.sleep(loadDesignTime)
        clearQueue(OpenROADOutputQueue)
        # Sheet0-5 are short DB testcases, so we set maxNewTokens to 512
        # Sheet6 is long Flow testcases, so we set maxNewTokens to 8192
        maxNewTokens = 512 if sheetIndex < 6 else 8192
        decoded = generate(model = model,
          tokenizer = tokenizer,
          prompt = prompt,
          pastKeyValues = DynamicCache(),
          temperature=0.3,
          topP=0.9,
          maxNewTokens=maxNewTokens
        )
        if modelName != "OpenROAD-Assistant/Script_Adaptor":
          code = decoded.split("```python")[-1].split("```")[0]
        else:
          code = decoded.split("<|begin_of_python|>")[-1].split("<|end_of_python|>")[0]
        codeTestSetIter[sheetIndex][passIndex + str(i + 2)] = code
        generatedCommand = processCodeString(code)
        # Send commands to OpenROAD and handle potential process termination
        while True:
          try:
            stdout, traceback = sendCommandOpenROAD(OpenROADProcess, generatedCommand, OpenROADOutputQueue, maxTestCaseWaitTime, commandFlushTime)
            break  # Exit loop if the command runs successfully
          except RuntimeError:
            print("OpenROADProcess terminated. Restarting the OpenROAD shell.")
            OpenROADProcess = runOpenROADShell(OpenROADPath, loadDesignTime, slaveOpenROAD, loadDesignType)  # Restart the OpenROAD process
            # Clear the output queues
            time.sleep(loadDesignTime)
            clearQueue(OpenROADOutputQueue)
        # Update the initial question based on output or errors
        stdout = stdout.encode('utf-8')
        stdout = re.sub(b'\x1b\].*?\n', b'', stdout)
        stdout = re.sub(b'\x1b\].*?\x07', b'', stdout)
        stdout = stdout.decode('utf-8')
        if passIndex == "C":
          codeTestSetIter[sheetIndex]["D"+str(i+2)] = str(stdout.strip())
        elif passIndex == "E":
          codeTestSetIter[sheetIndex]["F"+str(i+2)] = str(stdout.strip())
        elif passIndex == "G":
          codeTestSetIter[sheetIndex]["H"+str(i+2)] = str(stdout.strip())
        modelName_ = modelName.split("/")[-1] if modelName != "OpenROAD-Assistant/Script_Adaptor" else "ORA"
        codeTestSet.save(resultPath+modelName_+"-pass@K.xlsx")
        if not traceback:
          break

if __name__ == '__main__':
  parser = argparse.ArgumentParser(description = "parsing the path")
  parser.add_argument("--testSetPath", type = str, default = "../EDA-Corpus-v2/TestSet.xlsx")
  parser.add_argument("--resultPath", type = str, default = "../result/")
  parser.add_argument("--RAGApiPath", type = str, default = "../RAGData/RAGAPIs.csv")
  parser.add_argument("--RAGCodePath", type = str, default = "../RAGData/RAGCodePiece.csv")
  parser.add_argument('--OpenROADPath', type=str, default='../OpenROAD/build/src/openroad')
  parser.add_argument('--modelName', type=str, default='Qwen/Qwen2.5-Coder-32B-Instruct')
  parser.add_argument('--loadDesignTime', type=float, default=2)
  parser.add_argument('--maxTestCaseWaitTime', type=float, default=120)
  parser.add_argument('--commandFlushTime', type=float, default=0.1)
  pyargs = parser.parse_args()

  multiprocessing.set_start_method('spawn', force=True)
  masterOpenROAD, slaveOpenROAD = None, None
  masterLLM, slaveLLM = None, None
  OpenROADProcess = None
  
  try:
    # Set sleep count to prepare the process
    # Start the OpenROAD and LLM shell and keep it running
    masterOpenROAD, slaveOpenROAD = os.openpty()  # Create a pseudo-terminal

    OpenROADOutputQueue = queue.Queue()

    stopEvent = threading.Event()
    # Create threads to read both stdout and stderr into a single queue
    OpenROADStdoutThread = threading.Thread(
      target=readOpenROADOutput, 
      args=(masterOpenROAD, OpenROADOutputQueue, 'STDOUT', stopEvent)
    )
    
    OpenROADStdoutThread.daemon = True
    OpenROADStdoutThread.start()

    # Clear the process initialization messages from the queues
    while not OpenROADOutputQueue.empty():
      outputType, line = OpenROADOutputQueue.get_nowait()

    Run(testSetPath = pyargs.testSetPath,
        RAGApiPath = pyargs.RAGApiPath,
        RAGCodePath = pyargs.RAGCodePath,
        resultPath = pyargs.resultPath,
        modelName = pyargs.modelName,
        OpenROADPath = pyargs.OpenROADPath,
        loadDesignTime = pyargs.loadDesignTime,
        maxTestCaseWaitTime = pyargs.maxTestCaseWaitTime,
        commandFlushTime = pyargs.commandFlushTime
        )
  except KeyboardInterrupt:
    print("\nInterrupted. Terminating the Python shell process.")
